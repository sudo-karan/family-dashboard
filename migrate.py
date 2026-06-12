#!/usr/bin/env python3
"""One-time migration: old multi-sheet FD workbook -> flat CSVs for the new Google Sheet.

Usage:
    python3 migrate.py /path/to/Fixed_Deposits.xlsx [--outdir migrated]

Reads every per-account ledger sheet (skipping "Summary Sheet", "Consolidated Sheet"
and any "* Targets" sheet) and writes:

    <outdir>/FDs.csv       -> paste into the `FDs` tab of the new Google Sheet
    <outdir>/Accounts.csv  -> paste into the `Accounts` tab

Conversion rules (see README.md):
  - One row per deposit; `Account` = the old sheet/tab name, kept verbatim.
  - FD account numbers are treated as OPAQUE STRINGS (huge integers, leading
    zeros, "NA", "nsc ..." are all preserved verbatim, never parsed as numbers).
  - Dates are written as plain `YYYY-MM-DD` text.
  - `Tenure (Days)` is recomputed from the dates (blank when dates are missing).
  - Empty placeholder rows (Inactive with no principal and no dates) are skipped;
    Inactive rows that DO hold data are kept.
  - Defaults: Interest On Maturity = TRUE, Compounding = quarterly,
    Show In Dashboard = TRUE only for Active rows.
  - "Stable Money" keeps its per-FD bank name (column J) in `Remarks`, and maps
    to person "Jaskaran", bank "Stable Money".
"""

import argparse
import csv
import os
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

FD_HEADERS = [
    "ID", "Account", "Status", "FD Account Number", "Start Date", "End Date",
    "Tenure (Days)", "Interest %", "Principal Amount", "Maturity Amount",
    "Interest On Maturity", "Compounding", "Show In Dashboard", "Remarks",
]
ACCOUNT_HEADERS = ["Name", "Person", "Bank"]

SKIP_SHEETS = {"summary sheet", "consolidated sheet"}
# Banks whose display name differs from the raw token in the sheet name.
BANK_SPECIAL = {"indian": "Indian Bank", "axis": "Axis"}


def is_skipped_sheet(title):
    t = title.strip().lower()
    return t in SKIP_SHEETS or t.endswith("targets")


def person_bank(account_name):
    """Derive (person, bank) from an old sheet name like 'DAD HDFC'."""
    name = account_name.strip()
    if name.lower() == "stable money":
        return "Jaskaran", "Stable Money"
    parts = name.split()
    person = parts[0].title()
    rest = parts[1:]
    if not rest:
        return person, ""
    key = " ".join(rest).lower()
    if key in BANK_SPECIAL:
        return person, BANK_SPECIAL[key]
    words = []
    for w in rest:
        if w.isalpha() and w.isupper() and len(w) <= 5:
            words.append(w)            # acronyms: HDFC, SBI, ICICI ...
        elif w.islower():
            words.append(w.title())
        else:
            words.append(w)            # already mixed/title case: Post, Office
    return person, " ".join(words)


def opaque(v):
    """Account numbers etc.: preserve verbatim, never as parsed numbers."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        # xlsx stores numerics as doubles; render whole floats without sci-notation
        return "{:.0f}".format(v) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def iso_date(v, ctx, warnings):
    if v in (None, ""):
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d-%b-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    warnings.append("%s: could not parse date %r — kept verbatim, fix by hand" % (ctx, s))
    return s


def number(v, ctx, warnings):
    if v in (None, ""):
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("₹", "")
    if not s:
        return ""
    try:
        return float(s)
    except ValueError:
        warnings.append("%s: non-numeric amount %r — left blank" % (ctx, v))
        return ""


def fmt_num(v):
    """Render 50000.0 as '50000' and 7.1 as '7.1' for clean CSV cells."""
    if v == "" or v is None:
        return ""
    f = float(v)
    return "{:.0f}".format(f) if f.is_integer() else repr(f)


def tenure_days(start_iso, end_iso):
    try:
        s = datetime.strptime(start_iso, "%Y-%m-%d")
        e = datetime.strptime(end_iso, "%Y-%m-%d")
    except (ValueError, TypeError):
        return ""
    return (e - s).days


def find_header_row(ws):
    """Locate the ledger header row (col A == 'S.No'); data starts on the next row."""
    for r in range(1, min(ws.max_row, 6) + 1):
        a = ws.cell(row=r, column=1).value
        if a is not None and str(a).strip().lower() in ("s.no", "sno", "s no", "sr.no", "sr no"):
            return r
    return 3  # the family's workbook keeps headers on row 3, data from row 4


def convert(path, outdir):
    wb = openpyxl.load_workbook(path, data_only=True)
    fds, accounts, warnings = [], [], []

    for ws in wb.worksheets:
        title = ws.title.strip()
        if is_skipped_sheet(title):
            print("  skipping meta sheet: %r" % ws.title)
            continue

        person, bank = person_bank(title)
        accounts.append({"Name": title, "Person": person, "Bank": bank})

        hdr_row = find_header_row(ws)
        # Optional 10th column (Stable Money): per-FD bank name / remarks.
        hdr10 = ws.cell(row=hdr_row, column=10).value
        has_remarks_col = hdr10 is not None and str(hdr10).strip().lower() in ("remarks", "bank name")

        kept = 0
        for r in range(hdr_row + 1, ws.max_row + 1):
            ctx = "%s row %d" % (title, r)
            sno, status_raw, acct_no, sd, ed, _ten, rate, principal, maturity = (
                ws.cell(row=r, column=c).value for c in range(1, 10)
            )
            remarks = ws.cell(row=r, column=10).value if has_remarks_col else None

            row_vals = [sno, status_raw, acct_no, sd, ed, rate, principal, maturity, remarks]
            if all(v in (None, "") for v in row_vals):
                continue  # fully blank row

            status = str(status_raw).strip().title() if status_raw not in (None, "") else ""
            p_val = number(principal, ctx, warnings)
            m_val = number(maturity, ctx, warnings)
            start_iso = iso_date(sd, ctx, warnings)
            end_iso = iso_date(ed, ctx, warnings)

            has_principal = p_val != "" and p_val > 0
            has_dates = bool(start_iso or end_iso)
            if status != "Active" and not has_principal and not has_dates:
                continue  # empty placeholder row
            if status not in ("Active", "Inactive"):
                warnings.append("%s: status %r — defaulted to 'Active'" % (ctx, status_raw))
                status = "Active"

            fds.append({
                "Account": title,
                "Status": status,
                "FD Account Number": opaque(acct_no),
                "Start Date": start_iso,
                "End Date": end_iso,
                "Tenure (Days)": tenure_days(start_iso, end_iso),
                "Interest %": fmt_num(number(rate, ctx, warnings)),
                "Principal Amount": fmt_num(p_val),
                "Maturity Amount": fmt_num(m_val),
                "Interest On Maturity": "TRUE",
                "Compounding": "quarterly",
                "Show In Dashboard": "TRUE" if status == "Active" else "FALSE",
                "Remarks": opaque(remarks) if remarks not in (None, "") else "",
            })
            kept += 1
        print("  %-18s -> %2d FDs  (person=%s, bank=%s)" % (title, kept, person, bank))

    for i, fd in enumerate(fds, 1):
        fd["ID"] = i

    os.makedirs(outdir, exist_ok=True)
    fds_path = os.path.join(outdir, "FDs.csv")
    acc_path = os.path.join(outdir, "Accounts.csv")
    with open(fds_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FD_HEADERS)
        w.writeheader()
        w.writerows(fds)
    with open(acc_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=ACCOUNT_HEADERS)
        w.writeheader()
        w.writerows(accounts)

    # Totals over the dashboard set (Show In Dashboard = TRUE) — compare these
    # with the old workbook's own Summary Sheet to confirm nothing was lost.
    dash = [f for f in fds if f["Show In Dashboard"] == "TRUE"]
    tp = sum(float(f["Principal Amount"] or 0) for f in dash)
    tm = sum(float(f["Maturity Amount"] or 0) for f in dash)
    print("\nWrote %d FDs across %d accounts." % (len(fds), len(accounts)))
    print("  %s" % fds_path)
    print("  %s" % acc_path)
    print("\nDashboard set (Show In Dashboard = TRUE): %d FDs" % len(dash))
    print("  total principal : %s" % fmt_num(tp))
    print("  total maturity  : %s" % fmt_num(tm))
    print("Cross-check those two numbers against the old 'Summary Sheet' totals.")
    if warnings:
        print("\n%d warning(s):" % len(warnings))
        for msg in warnings:
            print("  ! %s" % msg)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", help="path to the old Fixed_Deposits.xlsx")
    ap.add_argument("--outdir", default="migrated", help="output directory (default: ./migrated)")
    args = ap.parse_args()
    print("Reading %s" % args.workbook)
    convert(args.workbook, args.outdir)


if __name__ == "__main__":
    main()
